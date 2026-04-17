# =============================================================================
# Script: gen_propuesta_asignacion.py
# Propósito: Excel con propuesta de reasignación de proyectos a PMs
#            con estimación de horas basada en datos reales de OpenAir
# PMO Operations - Blend360 Colombia | Abril 2026
# =============================================================================
import os, warnings
from datetime import datetime
from collections import defaultdict
warnings.filterwarnings('ignore')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.image import Image as XLImage

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

# ── PATHS ─────────────────────────────────────────────────────────────────────
BASE   = 'c:/Users/BOG-LAP-SER-176/Documents/PMO-Operations'
FOLDER = [f for f in os.listdir(BASE) if 'Asign' in f][0]
ROOT   = os.path.join(BASE, FOLDER)
INPUT  = os.path.join(ROOT, '1. Input')
OUTPUT = os.path.join(ROOT, '3. Output')
CHARTS = os.path.join(OUTPUT, 'charts_propuesta')
os.makedirs(CHARTS, exist_ok=True)

MAESTRO_FILE = os.path.join(INPUT, 'maestro.proyectos.xlsx')
HOURS_FILE   = os.path.join(INPUT, 'Horas marzo abril.xlsx')

# ── BRAND COLORS ──────────────────────────────────────────────────────────────
C_DARK   = '003059'
C_MID    = '0070C0'
C_LIGHT  = 'BDD7EE'
C_ORANGE = 'FF6B00'
C_GREEN  = '70AD47'
C_YELLOW = 'FFD966'
C_RED    = 'FF4D4D'
C_LGRAY  = 'F2F2F2'
C_GRAY   = '595959'
C_WHITE  = 'FFFFFF'
C_PURPLE = '7030A0'
C_TEAL   = '00B0F0'

PALETTE_HEX = ['#003059','#0070C0','#FF6B00','#70AD47','#FFD966',
               '#BDD7EE','#7030A0','#00B0F0','#FF4D4D','#A9D18E']

# ── PM ↔ OpenAir user mapping ─────────────────────────────────────────────────
PM_USER = {
    'Oscar Barragan':           'BARRAGAN, OSCAR',
    'Juan Bernal':              'BERNAL MORENO, JUAN CAMILO',
    'David Cortes':             'CORTES, DAVID',
    'Miguel Garcia':            'GARCIA, MIGUEL',
    'Kelly Carbonell':          'CARBONELL RAMOS, KELLY MARGARITA',
    'Daniel Sebastian Vargas':  'VARGAS CRISTANCHO, DANIEL SEBASTIAN',
    'Diana Castro':             'CASTRO, DIANA',
    'Diana Rojas':              'ROJAS CHARRY, DIANA CLEMENCIA',
    'Indira Duarte':            'DUARTE, INDIRA',
}
USER_PM = {v: k for k, v in PM_USER.items()}

# ── PROPOSED REASSIGNMENTS (project ID → new PM) ──────────────────────────────
# Miguel Garcia → ADRES Auditoría (takes from Indira Duarte)
# Superservicios → Diana Rojas (from Miguel Garcia)
# SED + MEN → Diana Castro (from Miguel Garcia)
# Hotel Planner → David Cortes (from Miguel Garcia)
# JPM → Juan Bernal (from David Cortes)
REASSIGN = {
    # ADR - AI Auditoría Médica (from Indira → Miguel)
    'P2457': 'Miguel Garcia', 'P2470': 'Miguel Garcia',
    'P2471': 'Miguel Garcia', 'P2472': 'Miguel Garcia',
    # SPD Superservicios (from Miguel → Diana Rojas)
    'P1966': 'Diana Rojas', 'P1967': 'Diana Rojas',
    'P2723': 'Diana Rojas', 'P2724': 'Diana Rojas',
    'P2725': 'Diana Rojas', 'P2726': 'Diana Rojas',
    # SED Secretaría de Educación (from Miguel → Diana Castro)
    'P2431': 'Diana Castro', 'P2561': 'Diana Castro',
    'P2562': 'Diana Castro', 'P2563': 'Diana Castro',
    # MEN Ministerio de Educación (from Miguel → Diana Castro)
    'P2611': 'Diana Castro', 'P2612': 'Diana Castro', 'P2634': 'Diana Castro',
    # HPL Hotel Planner (from Miguel → David Cortes)
    'P2671': 'David Cortes', 'P2786': 'David Cortes',
    # JPM (from David Cortes → Juan Bernal)
    'P1947': 'Juan Bernal', 'P1948': 'Juan Bernal',
}

# ── LOAD MAESTRO ──────────────────────────────────────────────────────────────
print("Cargando maestro de proyectos...")
wb_m = openpyxl.load_workbook(MAESTRO_FILE, data_only=True)
ws_m = wb_m['project']
maestro = list(ws_m.iter_rows(values_only=True))

active_projects = []   # list of dicts
for row in maestro[1:]:
    if row[8] != 'Activo': continue
    pid     = str(row[0]) if row[0] else ''
    sigla   = str(row[5]) if row[5] else ''
    obj     = str(row[3]) if row[3] else ''
    btype   = str(row[4]) if row[4] else ''
    tipo    = str(row[6]) if row[6] else ''
    cliente = str(row[7]) if row[7] else ''
    pm_orig = str(row[11]) if row[11] else 'N.A.'
    # Construct full name: Sigla - Objetivo - BlendType
    full_name = f"{sigla} - {obj} - {btype}"
    pm_prop = REASSIGN.get(pid, pm_orig)
    active_projects.append({
        'id': pid, 'sigla': sigla, 'objetivo': obj, 'btype': btype,
        'tipo': tipo, 'cliente': cliente, 'full_name': full_name,
        'pm_original': pm_orig, 'pm_propuesto': pm_prop,
        'reassigned': pid in REASSIGN,
    })

print(f"  Proyectos activos: {len(active_projects)}")

# ── LOAD HOURS (OpenAir) ──────────────────────────────────────────────────────
print("Cargando horas OpenAir...")
wb_h = openpyxl.load_workbook(HOURS_FILE)
ws_h = wb_h['openair (39)']
hours_rows = list(ws_h.iter_rows(values_only=True))

# pm_proj_hours[pm][openair_project] → March hours (PM's own time)
pm_proj_hours = defaultdict(lambda: defaultdict(float))
# total_proj_hours[openair_project] → total team hours March
total_proj_hours = defaultdict(float)
# pm_total_march[pm] → total PM hours March
pm_total_march = defaultdict(float)

for row in hours_rows[1:]:
    user    = str(row[1]) if row[1] else ''
    date    = row[0]
    hours   = float(row[9]) if row[9] else 0
    project = str(row[8]) if row[8] else ''

    if not isinstance(date, datetime): continue
    if date.year != 2026 or date.month != 3: continue   # March only (full month)

    total_proj_hours[project] += hours
    if user in USER_PM:
        pm = USER_PM[user]
        pm_proj_hours[pm][project] += hours
        pm_total_march[pm] += hours

# ── ESTIMATE PM HOURS PER MAESTRO PROJECT ─────────────────────────────────────
# Map: openair project keywords → maestro sigla
# We estimate PM management hours using the PM's own recorded hours per project
OA_MAP = {
    # sigla → list of OpenAir project name substrings to match
    'ADR':  ['adres', 'adres - ai', 'adres - mod', 'auditoria medica', 'adr -'],
    'SPD':  ['superserv', 'secretar', 'planeaci', 'repositorio', 'observatorio'],
    'SED':  ['secretary of educ', 'secretaria de educ', 'sed -'],
    'MEN':  ['ministerio de educ', 'colombia aprende', 'men -'],
    'HPL':  ['hotelplanner', 'hotel planner', 'hpl -'],
    'JPM':  ['jpm -', 'jpm cloud'],
    'BTG':  ['btg -', 'smartcash'],
    'CJU':  ['cju -', 'coljuegos', 'vigilancia y met', 'juegos localizados'],
    'ACA':  ['aca -', 'cali', 'data cali', 'virtualizaci'],
    'CAT':  ['cat -', 'agente cognitivo'],
    'CSJ':  ['csj -', 'continuidad operacion'],
    'SDH':  ['sdh -', 'secretaria del habitat'],
    'IGM':  ['igm -', 'ec - igm'],
    'MMN':  ['mmn -', 'metro medellin', 'metro medell', 'analitica de datos'],
    'ICF':  ['icf -', 'co - icfes', 'co - icf -', 'datalake'],
    'BID':  ['bid -'],
    'FDN':  ['fdn -'],
    'DIA':  ['dia -'],
    'PON':  ['pon -', 'iapol'],
    'FOA':  ['foa -', 'mojana'],
    'SOC':  ['soc -', 'supersociedades'],
    'UNA':  ['una -', 'unal -', 'co- universidad nacional'],
    'UCA':  ['ucaldas', 'universidad de caldas', 'co - universidad de caldas'],
    'UDS':  ['universidad sabana', 'co - universidad sabana'],
    'PRO':  ['pro -', 'progresion'],
    'CCC':  ['ccc -'],
    'SAN':  ['san -'],
    'AGR':  ['fiduagraria', 'co - fiduagraria'],
    'MTC':  ['mtc -', 'microcentros'],
}

def get_pm_hours_for_sigla(pm, sigla):
    """Return the PM's total March hours across all OpenAir projects matching sigla."""
    keywords = OA_MAP.get(sigla, [sigla.lower()])
    total = 0.0
    for oa_proj, h in pm_proj_hours[pm].items():
        oa_lower = oa_proj.lower()
        if any(k in oa_lower for k in keywords):
            total += h
    return total

def get_total_hours_for_sigla(sigla):
    """Return total team March hours for projects matching sigla."""
    keywords = OA_MAP.get(sigla, [sigla.lower()])
    total = 0.0
    for oa_proj, h in total_proj_hours.items():
        oa_lower = oa_proj.lower()
        if any(k in oa_lower for k in keywords):
            total += h
    return total

# ── GROUP PROJECTS BY (pm_propuesto, sigla, objetivo) for cleaner view ─────────
# Within each PM + client group, split PM hours proportionally across project IDs
from collections import OrderedDict

# Build groups: key = (pm_propuesto, sigla, objetivo) → list of project dicts
groups = defaultdict(list)
for p in active_projects:
    if p['pm_propuesto'] in ['N.A.', 'Sin Asignar', 'None', '']: continue
    key = (p['pm_propuesto'], p['sigla'], p['objetivo'])
    groups[key].append(p)

# For each group, calculate estimated PM hours
PM_MIN_HOURS = {
    'Implementaci\u00f3n': 12, 'Operaci\u00f3n': 10, 'Software SW': 8,
    'Data': 6, 'Interno': 4,
}

# ── Compute pm_hours_per_sigla ONCE per (pm, sigla) then divide across groups ─
# This avoids double-counting when a sigla has multiple objective groups
pm_sigla_keys = defaultdict(int)   # (pm, sigla) → count of groups
for (pm, sigla, obj) in groups:
    pm_sigla_keys[(pm, sigla)] += 1

pm_sigla_hours_cache = {}  # (pm, sigla) → total pm hours for sigla

for (pm, sigla, obj), projs in groups.items():
    key = (pm, sigla)
    if key not in pm_sigla_hours_cache:
        h = get_pm_hours_for_sigla(pm, sigla)
        # If reassigned and no hours yet, use previous PM's hours
        if h == 0 and projs[0]['reassigned']:
            prev_pm = projs[0]['pm_original']
            h = get_pm_hours_for_sigla(prev_pm, sigla)
        # If still 0, estimate from project types
        if h == 0:
            tipo_sum = sum(PM_MIN_HOURS.get(p['tipo'], 8) for p in projs)
            h = tipo_sum
        pm_sigla_hours_cache[key] = h

enriched = []   # final list for Excel rows
for (pm, sigla, obj), projs in groups.items():
    key = (pm, sigla)
    n_groups = pm_sigla_keys[key]
    # Spread sigla hours evenly across groups for that sigla
    pm_hours_sigla = pm_sigla_hours_cache[key] / n_groups

    team_hours_sigla = get_total_hours_for_sigla(sigla)
    if team_hours_sigla == 0:
        team_hours_sigla = pm_hours_sigla * 4

    # Distribute evenly across sub-projects in each group
    hrs_per_proj   = pm_hours_sigla / len(projs)
    team_per_proj  = team_hours_sigla / (n_groups * len(projs))

    for p in projs:
        enriched.append({
            **p,
            'pm_hours_est':    round(hrs_per_proj, 1),
            'team_hours_est':  round(team_per_proj, 1),
            'pct_pm_capacity': round(hrs_per_proj / 193.6 * 100, 1),
        })

# Sort: pm_propuesto → sigla → objetivo
enriched.sort(key=lambda x: (x['pm_propuesto'], x['sigla'], x['objetivo']))

# Summary per PM (proposed)
pm_summary = defaultdict(lambda: {
    'n_projects': 0, 'pm_hours': 0.0, 'team_hours': 0.0, 'reassigned': 0
})
for p in enriched:
    pm = p['pm_propuesto']
    pm_summary[pm]['n_projects'] += 1
    pm_summary[pm]['pm_hours']   += p['pm_hours_est']
    pm_summary[pm]['team_hours'] += p['team_hours_est']
    if p['reassigned']:
        pm_summary[pm]['reassigned'] += 1

PM_ORDER = sorted(pm_summary.keys())
print(f"\nResumen propuesta ({len(enriched)} sub-proyectos, {len(groups)} agrupaciones):")
for pm in PM_ORDER:
    s = pm_summary[pm]
    pct = s['pm_hours'] / 193.6 * 100
    print(f"  {pm}: {s['n_projects']} proyectos | ~{s['pm_hours']:.0f}h PM est. ({pct:.0f}%) | {s['reassigned']} reasignados")

# ── GENERATE CHARTS ───────────────────────────────────────────────────────────
print("\nGenerando gráficas...")

def spine(ax):
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#CCCCCC')
    ax.spines['bottom'].set_color('#CCCCCC')
    ax.tick_params(colors='#404040', labelsize=8)
    ax.yaxis.grid(True, linestyle='--', alpha=0.4, color='#DDDDDD')
    ax.set_axisbelow(True)

CAPACITY = 193.6

# ── Chart 1: Antes vs Después – horas PM estimadas por PM ────────────────────
# "Before": use actual March hours from OpenAir per PM
before = {pm: pm_total_march.get(PM_USER.get(pm,''), 0) for pm in PM_ORDER}
after  = defaultdict(float)
for p in enriched:
    after[p['pm_propuesto']] += p['pm_hours_est']

fig, axes = plt.subplots(1, 2, figsize=(13, 5), sharey=False)
pms = PM_ORDER
x = np.arange(len(pms))
w = 0.55

# Before
ax = axes[0]
vals_b = [before.get(pm, 0) for pm in pms]
colors_b = ['#C9504C' if v > CAPACITY else '#FFD966' if v > CAPACITY*0.9
             else '#003059' for v in vals_b]
bars = ax.bar(x, vals_b, w, color=colors_b, edgecolor='white')
ax.axhline(CAPACITY, color='#FF6B00', lw=1.5, ls='--', label='Capacidad (193.6h)')
ax.set_xticks(x); ax.set_xticklabels([p.replace(' ', '\n') for p in pms], fontsize=7)
ax.set_title('Distribución ACTUAL\n(Marzo 2026)', fontsize=10, fontweight='bold', color='#003059')
ax.set_ylabel('Horas PM estimadas/mes', fontsize=8)
ax.legend(fontsize=7); spine(ax)
for bar, v in zip(bars, vals_b):
    if v > 0: ax.text(bar.get_x()+bar.get_width()/2, v+2, f'{v:.0f}h',
                      ha='center', va='bottom', fontsize=7)

# After
ax = axes[1]
vals_a = [after.get(pm, 0) for pm in pms]
colors_a = ['#C9504C' if v > CAPACITY else '#FFD966' if v > CAPACITY*0.9
             else '#70AD47' for v in vals_a]
bars2 = ax.bar(x, vals_a, w, color=colors_a, edgecolor='white')
ax.axhline(CAPACITY, color='#FF6B00', lw=1.5, ls='--', label='Capacidad (193.6h)')
ax.set_xticks(x); ax.set_xticklabels([p.replace(' ', '\n') for p in pms], fontsize=7)
ax.set_title('Distribución PROPUESTA\n(Reasignaciones aplicadas)', fontsize=10,
             fontweight='bold', color='#003059')
ax.legend(fontsize=7); spine(ax)
for bar, v in zip(bars2, vals_a):
    if v > 0: ax.text(bar.get_x()+bar.get_width()/2, v+2, f'{v:.0f}h',
                      ha='center', va='bottom', fontsize=7)

fig.suptitle('Horas de Gestión PM Estimadas por Mes – Antes vs Después de Reasignaciones',
             fontsize=11, fontweight='bold', color='#003059', y=1.01)
plt.tight_layout()
c1 = os.path.join(CHARTS, 'c1_antes_despues.png')
plt.savefig(c1, dpi=150, bbox_inches='tight', facecolor='white'); plt.close()
print(f"  Gráfica 1: {c1}")

# ── Chart 2: Número de proyectos antes vs después ─────────────────────────────
n_before = defaultdict(int)
n_after  = defaultdict(int)
for p in enriched:
    n_after[p['pm_propuesto']]  += 1
    n_before[p['pm_original']] += 1

fig, ax = plt.subplots(figsize=(10, 5))
x = np.arange(len(pms))
w = 0.35
b1 = ax.bar(x - w/2, [n_before.get(pm,0) for pm in pms], w,
            label='Actual', color='#003059', alpha=0.85)
b2 = ax.bar(x + w/2, [n_after.get(pm,0) for pm in pms],  w,
            label='Propuesto', color='#0070C0', alpha=0.85)
ax.set_xticks(x); ax.set_xticklabels([p.replace(' ', '\n') for p in pms], fontsize=8)
ax.set_ylabel('Número de Sub-Proyectos', fontsize=9)
ax.set_title('Proyectos Asignados por PM – Actual vs Propuesto',
             fontsize=11, fontweight='bold', color='#003059')
ax.legend(fontsize=9); spine(ax)
for bar in list(b1)+list(b2):
    v = bar.get_height()
    if v > 0: ax.text(bar.get_x()+bar.get_width()/2, v+0.3, str(int(v)),
                      ha='center', va='bottom', fontsize=7)
plt.tight_layout()
c2 = os.path.join(CHARTS, 'c2_n_proyectos.png')
plt.savefig(c2, dpi=150, bbox_inches='tight', facecolor='white'); plt.close()
print(f"  Gráfica 2: {c2}")

# ── Chart 3: % Utilización propuesta por PM ───────────────────────────────────
fig, ax = plt.subplots(figsize=(9, 4.5))
pcts = [after.get(pm,0)/CAPACITY*100 for pm in pms]
colors_p = ['#C9504C' if p > 100 else '#FFD966' if p > 85 else '#70AD47' for p in pcts]
bars = ax.barh(pms, pcts, color=colors_p, edgecolor='white', height=0.6)
ax.axvline(100, color='#FF6B00', ls='--', lw=1.5, label='100% capacidad')
ax.axvline(85,  color='#FFD966', ls=':',  lw=1.2, label='85% umbral óptimo')
ax.set_xlabel('% Utilización estimada', fontsize=9)
ax.set_title('% Utilización por PM – Propuesta de Asignación\n(Base: 193.6h / mes)',
             fontsize=10, fontweight='bold', color='#003059')
ax.legend(fontsize=8, loc='lower right')
spine(ax); ax.spines['left'].set_visible(True)
for bar, pct in zip(bars, pcts):
    ax.text(pct+0.5, bar.get_y()+bar.get_height()/2,
            f'{pct:.1f}%', va='center', fontsize=8)
ax.set_xlim(0, 115)
plt.tight_layout()
c3 = os.path.join(CHARTS, 'c3_utilizacion_propuesta.png')
plt.savefig(c3, dpi=150, bbox_inches='tight', facecolor='white'); plt.close()
print(f"  Gráfica 3: {c3}")

# ── Chart 4: Stacked bar – top clients per PM (proposed) ─────────────────────
# Get top siglas per PM for stacked view
pm_sigla_hours = defaultdict(lambda: defaultdict(float))
for p in enriched:
    pm_sigla_hours[p['pm_propuesto']][p['sigla']] += p['pm_hours_est']

all_siglas_sorted = sorted(
    set(p['sigla'] for p in enriched),
    key=lambda s: -sum(pm_sigla_hours[pm].get(s,0) for pm in PM_ORDER)
)[:12]   # top 12 siglas by total hours

fig, ax = plt.subplots(figsize=(13, 6))
x = np.arange(len(pms))
w = 0.6
bottoms = np.zeros(len(pms))
for i, sig in enumerate(all_siglas_sorted):
    vals = [pm_sigla_hours[pm].get(sig, 0) for pm in pms]
    if sum(vals) == 0: continue
    color = PALETTE_HEX[i % len(PALETTE_HEX)]
    ax.bar(x, vals, w, bottom=bottoms, label=sig, color=color, edgecolor='white')
    bottoms += np.array(vals)

ax.axhline(CAPACITY, color='#FF6B00', lw=1.5, ls='--', label='Capacidad')
ax.set_xticks(x); ax.set_xticklabels([p.replace(' ','\n') for p in pms], fontsize=8)
ax.set_ylabel('Horas PM estimadas / mes', fontsize=9)
ax.set_title('Distribución de Horas PM por Cliente – Propuesta\n(Apilado por Sigla de Cliente)',
             fontsize=10, fontweight='bold', color='#003059')
ax.legend(fontsize=7, bbox_to_anchor=(1.01, 1), loc='upper left', ncol=1)
spine(ax)
plt.tight_layout()
c4 = os.path.join(CHARTS, 'c4_clientes_apilado.png')
plt.savefig(c4, dpi=150, bbox_inches='tight', facecolor='white'); plt.close()
print(f"  Gráfica 4: {c4}")

# ── Chart 5: Proyectos reasignados (highlight) ────────────────────────────────
reassigned_rows = [p for p in enriched if p['reassigned']]
# Group by objetivo+sigla for display
seen = {}
for p in reassigned_rows:
    k = (p['sigla'], p['objetivo'][:40])
    if k not in seen:
        seen[k] = {'sigla': p['sigla'], 'obj': p['objetivo'][:40],
                   'pm_ant': p['pm_original'], 'pm_nue': p['pm_propuesto'],
                   'hours': p['pm_hours_est']}
    else:
        seen[k]['hours'] += p['pm_hours_est']

rs = sorted(seen.values(), key=lambda x: -x['hours'])
labels   = [f"{r['sigla']} - {r['obj'][:35]}" for r in rs]
hours_rs = [r['hours'] for r in rs]
pm_antes = [r['pm_ant'] for r in rs]
pm_nuevo = [r['pm_nue'] for r in rs]

pm_colors = {pm: PALETTE_HEX[i] for i, pm in enumerate(PM_ORDER)}

fig, ax = plt.subplots(figsize=(11, max(4, len(rs)*0.55 + 1)))
colors_rs = [pm_colors.get(p,'#595959') for p in pm_nuevo]
bars = ax.barh(labels, hours_rs, color=colors_rs, edgecolor='white', height=0.6)
ax.set_xlabel('Horas PM estimadas / mes', fontsize=9)
ax.set_title('Proyectos Reasignados – Horas de Gestión Estimadas\n(Color = PM propuesto)',
             fontsize=10, fontweight='bold', color='#003059')
spine(ax); ax.spines['left'].set_visible(True)
for bar, v, pa, pn in zip(bars, hours_rs, pm_antes, pm_nuevo):
    ax.text(v+0.5, bar.get_y()+bar.get_height()/2,
            f'{v:.0f}h  |  {pa[:15]} → {pn[:15]}',
            va='center', fontsize=7, color='#404040')
# Legend
handles = [mpatches.Patch(color=pm_colors.get(pm,'#595959'), label=pm)
           for pm in PM_ORDER if pm in set(pm_nuevo)]
ax.legend(handles=handles, fontsize=8, loc='lower right')
plt.tight_layout()
c5 = os.path.join(CHARTS, 'c5_reasignados.png')
plt.savefig(c5, dpi=150, bbox_inches='tight', facecolor='white'); plt.close()
print(f"  Gráfica 5: {c5}")

print("Gráficas completadas.")

# ── BUILD EXCEL ───────────────────────────────────────────────────────────────
print("\nGenerando Excel...")
wb = openpyxl.Workbook()

# ── Styles ────────────────────────────────────────────────────────────────────
thin  = Side(style='thin',   color='CCCCCC')
thick = Side(style='medium', color='003059')
bd    = Border(left=thin, right=thin, top=thin, bottom=thin)
bd_hd = Border(left=thick, right=thick, top=thick, bottom=thick)

def fill(c):   return PatternFill('solid', fgColor=c)
def fnt(sz=10, bold=False, color='000000', italic=False):
    return Font(name='Calibri', size=sz, bold=bold, color=color, italic=italic)
def aln(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr_cell(ws, row, col, val, width=None, fill_c=C_DARK, sz=10, wrap=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = fnt(sz, bold=True, color=C_WHITE)
    c.fill = fill(fill_c)
    c.alignment = aln(wrap=wrap)
    c.border = bd
    if width: ws.column_dimensions[get_column_letter(col)].width = width
    return c

def dat_cell(ws, row, col, val, bold=False, bg=None, fc='000000', fmt=None,
             halign='center', wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = fnt(10, bold=bold, color=fc)
    c.border = bd
    c.alignment = aln(halign, wrap=wrap)
    if bg: c.fill = fill(bg)
    if fmt: c.number_format = fmt
    return c

def set_bg(cell, hex_color):
    cell.fill = fill(hex_color)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – PROPUESTA DE ASIGNACIÓN (detalle por proyecto)
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Propuesta de Asignación'
ws1.sheet_view.showGridLines = False

# Title block
ws1.merge_cells('A1:L1')
ws1['A1'] = 'PROPUESTA DE REASIGNACIÓN DE PROYECTOS – BLEND360 COLOMBIA'
ws1['A1'].font = fnt(14, bold=True, color=C_WHITE)
ws1['A1'].fill = fill(C_DARK)
ws1['A1'].alignment = aln(wrap=False)
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:L2')
ws1['A2'] = 'Estimación de horas de gestión PM  |  Base: datos reales OpenAir Marzo 2026  |  Capacidad: 193.6 h/mes (44h/sem × 4.4 sem)'
ws1['A2'].font = fnt(9, italic=True, color=C_WHITE)
ws1['A2'].fill = fill(C_MID)
ws1['A2'].alignment = aln(wrap=False)
ws1.row_dimensions[2].height = 18

# Headers row 4
col_defs = [
    ('PM Propuesto',              20),
    ('ID Proyecto',               10),
    ('Sigla',                      8),
    ('Nombre Completo Proyecto',  52),
    ('Cliente',                   28),
    ('Tipo Proyecto',             16),
    ('Blend Type',                12),
    ('Horas PM\nest./mes',        11),
    ('% Capacidad\nPM',          11),
    ('Horas Equipo\nest./mes',    12),
    ('Estado\nReasignación',      14),
    ('PM Original',               20),
]
ws1.row_dimensions[4].height = 32
for ci, (hdr, wid) in enumerate(col_defs, start=1):
    hdr_cell(ws1, 4, ci, hdr, width=wid, fill_c=C_DARK)

# Data rows
pm_color_map = {
    'Oscar Barragan':           C_DARK,
    'Juan Bernal':              C_MID,
    'David Cortes':             'FF6B00',
    'Miguel Garcia':            '7030A0',
    'Kelly Carbonell':          '00B0F0',
    'Daniel Sebastian Vargas':  '70AD47',
    'Diana Castro':             'C9504C',
    'Diana Rojas':              'A9D18E',
    'Indira Duarte':            '595959',
}

row_i = 5
prev_pm = ''
for p in enriched:
    pm   = p['pm_propuesto']
    even = (row_i % 2 == 0)
    row_bg = C_LGRAY if even else C_WHITE

    # PM group header
    if pm != prev_pm:
        ws1.merge_cells(f'A{row_i}:L{row_i}')
        c = ws1.cell(row=row_i, column=1,
                     value=f'▶  {pm}  –  {pm_summary[pm]["n_projects"]} sub-proyectos  |  '
                           f'~{pm_summary[pm]["pm_hours"]:.0f}h est./mes  '
                           f'({pm_summary[pm]["pm_hours"]/CAPACITY*100:.0f}% capacidad)')
        c.font = fnt(11, bold=True, color=C_WHITE)
        c.fill = fill(pm_color_map.get(pm, C_DARK))
        c.alignment = aln('left')
        ws1.row_dimensions[row_i].height = 22
        row_i += 1
        prev_pm = pm

    # Project row
    reasign_flag  = '🔄 Reasignado' if p['reassigned'] else '—'
    reasign_color = 'FFD966' if p['reassigned'] else row_bg

    dat_cell(ws1, row_i, 1,  pm,                bg=row_bg, halign='left')
    dat_cell(ws1, row_i, 2,  p['id'],           bg=row_bg)
    dat_cell(ws1, row_i, 3,  p['sigla'],        bg=row_bg)
    dat_cell(ws1, row_i, 4,  p['full_name'],    bg=row_bg, halign='left', wrap=True)
    dat_cell(ws1, row_i, 5,  p['cliente'],      bg=row_bg, halign='left', wrap=True)
    dat_cell(ws1, row_i, 6,  p['tipo'],         bg=row_bg)
    dat_cell(ws1, row_i, 7,  p['btype'],        bg=row_bg)
    dat_cell(ws1, row_i, 8,  p['pm_hours_est'], bg=row_bg, fmt='#,##0.0')
    # % capacity with color
    pct_val = p['pct_pm_capacity'] / 100
    c_pct = ws1.cell(row=row_i, column=9, value=pct_val)
    c_pct.font    = fnt(10)
    c_pct.border  = bd
    c_pct.alignment = aln()
    c_pct.number_format = '0.0%'
    pct_bg = 'FF4D4D' if p['pct_pm_capacity']>15 else C_YELLOW if p['pct_pm_capacity']>8 else C_GREEN
    c_pct.fill = fill(pct_bg)
    if pct_bg in ['FF4D4D', C_GREEN]: c_pct.font = fnt(10, color=C_WHITE)

    dat_cell(ws1, row_i, 10, p['team_hours_est'], bg=row_bg, fmt='#,##0.0')
    dat_cell(ws1, row_i, 11, reasign_flag,
             bg=reasign_color, bold=p['reassigned'],
             fc=C_DARK if p['reassigned'] else C_GRAY)
    dat_cell(ws1, row_i, 12, p['pm_original'], bg=row_bg, halign='left')

    ws1.row_dimensions[row_i].height = 18
    row_i += 1

ws1.freeze_panes = 'A5'
ws1.row_dimensions[3].height = 8  # spacer

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – RESUMEN POR PM
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Resumen por PM')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:I1')
ws2['A1'] = 'RESUMEN DE CARGA POR PM – PROPUESTA DE REASIGNACIÓN'
ws2['A1'].font = fnt(13, bold=True, color=C_WHITE)
ws2['A1'].fill = fill(C_DARK)
ws2['A1'].alignment = aln()
ws2.row_dimensions[1].height = 28

hdrs2 = ['Project Manager', 'Proyectos\nActuales', 'Proyectos\nPropuestos',
         'Δ Proyectos', 'Horas PM\nest./mes', '% Capacidad\n(193.6h/mes)',
         'Estado Carga', 'Proyectos\nReasignados', 'Siglas de Proyectos']
widths2 = [26, 12, 12, 10, 13, 14, 16, 14, 45]
ws2.row_dimensions[3].height = 32
for ci, (h, w) in enumerate(zip(hdrs2, widths2), start=1):
    hdr_cell(ws2, 3, ci, h, width=w, fill_c=C_MID)

# Count actual projects per PM (original assignments)
n_actual = defaultdict(int)
for p in enriched:
    n_actual[p['pm_original']] += 1

for ri, pm in enumerate(PM_ORDER, start=4):
    s   = pm_summary[pm]
    n_a = n_actual.get(pm, 0)
    n_p = s['n_projects']
    delta = n_p - n_a
    pct = s['pm_hours'] / CAPACITY * 100

    if pct > 100:   estado = 'Sobrecargado ⚠'; est_bg = C_RED;    est_fc = C_WHITE
    elif pct > 85:  estado = 'Carga Alta';     est_bg = C_YELLOW; est_fc = '595959'
    elif pct > 60:  estado = 'Carga Normal';   est_bg = C_GREEN;  est_fc = C_WHITE
    else:           estado = 'Carga Baja';     est_bg = C_LIGHT;  est_fc = '003059'

    bg = C_LGRAY if ri % 2 == 0 else C_WHITE
    # Siglas list
    siglas_set = sorted(set(p['sigla'] for p in enriched if p['pm_propuesto'] == pm))

    dat_cell(ws2, ri, 1, pm,      bg=bg, halign='left', bold=True)
    dat_cell(ws2, ri, 2, n_a,     bg=bg)
    dat_cell(ws2, ri, 3, n_p,     bg=bg)
    dc = ws2.cell(row=ri, column=4, value=delta)
    dc.font = fnt(10, bold=True,
                  color='C9504C' if delta > 0 else '70AD47' if delta < 0 else '595959')
    dc.fill = fill(bg); dc.border = bd; dc.alignment = aln()
    dat_cell(ws2, ri, 5, s['pm_hours'],   bg=bg, fmt='#,##0.0')
    pct_c = ws2.cell(row=ri, column=6, value=pct/100)
    pct_c.number_format = '0.0%'; pct_c.border = bd; pct_c.alignment = aln()
    pct_c.fill = fill('FF4D4D' if pct>100 else C_YELLOW if pct>85 else C_GREEN)
    pct_c.font = fnt(10, bold=True, color=C_WHITE if pct>85 else '595959')
    c_est = ws2.cell(row=ri, column=7, value=estado)
    c_est.font = fnt(10, bold=True, color=est_fc); c_est.fill = fill(est_bg)
    c_est.border = bd; c_est.alignment = aln()
    dat_cell(ws2, ri, 8, s['reassigned'], bg=bg,
             fc=C_ORANGE if s['reassigned'] > 0 else '595959',
             bold=s['reassigned'] > 0)
    dat_cell(ws2, ri, 9, ' | '.join(siglas_set), bg=bg, halign='left', wrap=True)
    ws2.row_dimensions[ri].height = 22

# Totals
r_tot = len(PM_ORDER) + 4
ws2.merge_cells(f'A{r_tot}:D{r_tot}')
ws2[f'A{r_tot}'] = 'TOTAL'
ws2[f'A{r_tot}'].font = fnt(10, bold=True, color=C_WHITE)
ws2[f'A{r_tot}'].fill = fill(C_DARK)
ws2[f'A{r_tot}'].alignment = aln()
tot_h = sum(pm_summary[pm]['pm_hours'] for pm in PM_ORDER)
dat_cell(ws2, r_tot, 5, tot_h, bg=C_DARK, fmt='#,##0.0',
         bold=True, fc=C_WHITE)
ws2.row_dimensions[r_tot].height = 22
ws2.freeze_panes = 'A4'

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – CAMBIOS DE REASIGNACIÓN
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Cambios de Reasignación')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:G1')
ws3['A1'] = 'DETALLE DE CAMBIOS EN ASIGNACIÓN DE PROYECTOS'
ws3['A1'].font = fnt(13, bold=True, color=C_WHITE)
ws3['A1'].fill = fill(C_DARK); ws3['A1'].alignment = aln()
ws3.row_dimensions[1].height = 28

ws3.merge_cells('A2:G2')
ws3['A2'] = ('Los siguientes proyectos cambian de PM. Las horas estimadas se basan en el '
             'tiempo real registrado por el PM anterior en OpenAir (Marzo 2026).')
ws3['A2'].font = fnt(9, italic=True, color=C_WHITE)
ws3['A2'].fill = fill(C_MID); ws3['A2'].alignment = aln('left')
ws3.row_dimensions[2].height = 16

hdrs3 = ['ID Proyecto', 'Nombre Completo Proyecto', 'Cliente',
         'PM Anterior', 'PM Nuevo', 'Horas PM est./mes', 'Motivo del Cambio']
widths3 = [11, 55, 28, 24, 24, 16, 40]
ws3.row_dimensions[4].height = 30
for ci, (h, w) in enumerate(zip(hdrs3, widths3), start=1):
    hdr_cell(ws3, 4, ci, h, width=w, fill_c=C_MID)

motivos = {
    'Miguel Garcia':    'Reasignación para balancear carga y alinear especialidad',
    'Diana Rojas':      'Agrupación de proyectos de datos/observatorios',
    'Diana Castro':     'Alineación con portafolio Educación',
    'David Cortes':     'Alineación con portafolio infraestructura cloud',
    'Juan Bernal':      'Alineación con portafolio cloud managed services',
}

reassign_rows = sorted([p for p in enriched if p['reassigned']],
                        key=lambda x: (x['pm_propuesto'], x['sigla']))
ri = 5
for p in reassign_rows:
    bg = C_LGRAY if ri % 2 == 0 else 'FFFCE8'
    dat_cell(ws3, ri, 1, p['id'],         bg=bg)
    dat_cell(ws3, ri, 2, p['full_name'],  bg=bg, halign='left', wrap=True)
    dat_cell(ws3, ri, 3, p['cliente'],    bg=bg, halign='left', wrap=True)
    c_ant = ws3.cell(row=ri, column=4, value=p['pm_original'])
    c_ant.fill = fill(bg); c_ant.border = bd; c_ant.alignment = aln()
    c_ant.font = fnt(10, color='C9504C')
    c_nue = ws3.cell(row=ri, column=5, value=p['pm_propuesto'])
    c_nue.fill = fill(bg); c_nue.border = bd; c_nue.alignment = aln()
    c_nue.font = fnt(10, bold=True, color='70AD47')
    dat_cell(ws3, ri, 6, p['pm_hours_est'], bg=bg, fmt='#,##0.0')
    dat_cell(ws3, ri, 7, motivos.get(p['pm_propuesto'], '—'), bg=bg,
             halign='left', wrap=True)
    ws3.row_dimensions[ri].height = 22
    ri += 1

ws3.freeze_panes = 'A5'

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – GRÁFICAS
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Gráficas')
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:N1')
ws4['A1'] = 'VISUALIZACIÓN – PROPUESTA DE REASIGNACIÓN DE PROYECTOS'
ws4['A1'].font = fnt(13, bold=True, color=C_WHITE)
ws4['A1'].fill = fill(C_DARK); ws4['A1'].alignment = aln()
ws4.row_dimensions[1].height = 28

# Place charts in a single column layout (A) to avoid merge conflicts
chart_configs = [
    (c1, 2,  'Antes vs Después: Horas de gestión PM por mes',   0.72),
    (c2, 30, 'Número de proyectos: Actual vs Propuesto',         0.72),
    (c3, 58, 'Utilización de capacidad – Propuesta',             0.60),
    (c4, 82, 'Distribución de horas por cliente (apilado)',      0.82),
    (c5, 116,'Proyectos reasignados – Horas estimadas',          0.65),
]

ws4.column_dimensions['A'].width = 2
for img_path, start_row, title, scale in chart_configs:
    r = start_row
    # Title: write value before merging
    tc = ws4.cell(row=r, column=1, value=title)
    tc.font = fnt(10, bold=True, color=C_WHITE)
    tc.fill = fill(C_MID); tc.alignment = aln('left')
    ws4.row_dimensions[r].height = 18
    try:
        ws4.merge_cells(f'A{r}:P{r}')
    except Exception:
        pass
    try:
        img = XLImage(img_path)
        img.width  = int(img.width  * scale)
        img.height = int(img.height * scale)
        ws4.add_image(img, f'A{r+1}')
    except Exception as e:
        print(f"  Warning: no se pudo insertar imagen {img_path}: {e}")

# ═══════════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════════
out_path = os.path.join(OUTPUT, 'Propuesta_Reasignacion_PM_Blend360_new.xlsx')
wb.save(out_path)
print(f"\nExcel guardado: {out_path}")

# ── Copy script to Process folder ─────────────────────────────────────────────
import shutil
dest = os.path.join(ROOT, '2. Process', 'gen_propuesta_asignacion.py')
shutil.copy2(__file__, dest)

print("\n" + "="*60)
print("COMPLETADO")
print("="*60)
print(f"  {out_path}")
